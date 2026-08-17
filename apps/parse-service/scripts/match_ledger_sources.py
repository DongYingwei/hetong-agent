#!/usr/bin/env python3
"""只读：按金额优先、甲乙方次之匹配 Markdown 来源与审核台账。"""
from __future__ import annotations
import csv,json,pathlib,re
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from openpyxl import load_workbook

ROOT=pathlib.Path('/home/wdy/data/code/hetong-agent'); MD=ROOT/'data/md-file'; XLSX=ROOT/'demo/合同台账-V2.xlsx'
OUT=MD/'ledger-source-match.csv'
# 人工确认优先于自动规则；键为来源相对路径的唯一片段。
OVERRIDES={
 'CM-2025077 中国移动通信集团内蒙古有限公司':'HSKJ/C-CM-2025077',
 '业务-RJ-2025178-104-零跑汽车':'HSKJ/C-RJ-2025146',
}
def norm(v): return re.sub(r'[^0-9A-Za-z\u4e00-\u9fff]','',str(v or '')).lower()
def amount(v):
 try: return Decimal(str(v).replace(',','').strip())
 except (InvalidOperation,ValueError): return None
def md_amounts(text):
 vals=[]
 for x in re.findall(r'(?<![\d.])\d{3,}(?:,\d{3})*(?:\.\d+)?',text):
  v=amount(x)
  if v and v>=1000: vals.append(v)
 return set(vals)
def party(text, labels):
 for line in text.splitlines()[:300]:
  if any(x in line for x in labels):
   v=re.split(r'[:：]',line,1)
   if len(v)>1 and len(norm(v[1]))>3: return norm(v[1])
 return ''
def code(v):
 m=re.search(r'(?:CM|RJ|QC|QT|DL|HJ)-?(\d{7})',str(v),re.I); return m.group(1) if m else ''
ws=load_workbook(XLSX,read_only=True,data_only=True)['合同台账']; h=[str(x or '').strip() for x in next(ws.iter_rows(min_row=3,max_row=3,values_only=True))]
ledger=[dict(zip(h,r)) for r in ws.iter_rows(min_row=5,values_only=True) if any(r)]
manifest=json.loads((MD/'manifest.json').read_text())
report=[]
for e in manifest['entries']:
 for src in e['sources']:
  text=(MD/e['markdown_file']).read_text(encoding='utf8'); amts=md_amounts(text); a=party(text,['甲方','委托方']); b=party(text,['乙方','受托方']); sc=[]
  for row in ledger:
   money=amount(row.get('合同金额（含税）')); money_hit=bool(money and money in amts)
   pa=norm(row.get('客户名称')); party_hit=bool(pa and (pa in a or a in pa))
   title=SequenceMatcher(None,norm(src['relative_path']),norm(row.get('合同名称'))).ratio()
   # 文件名中的编号优先于父文件夹编号（一个文件夹可含 RJ-2025146、RJ-2025178 两合同）。
   c=code(pathlib.PurePosixPath(src['relative_path']).name) or code(src['relative_path']); code_hit=bool(c and c in norm(row.get('合同号')))
   score=(100 if money_hit else 0)+(20 if party_hit else 0)+(10 if code_hit else 0)+title
   sc.append((score,money_hit,party_hit,code_hit,row))
  sc.sort(key=lambda x:x[0],reverse=True); best=sc[0]; second=sc[1]
  status='confirmed' if ((best[1] and (best[2] or best[3]) and best[0]-second[0]>=10) or (best[3] and best[0]-second[0]>=5)) else 'needs_review'
  forced=next((no for hint,no in OVERRIDES.items() if hint in src['relative_path']),None)
  if forced:
   row=next(r for r in ledger if r['合同号']==forced)
   best=(best[0],False,False,False,row); status='confirmed_user'
  report.append({'source':src['relative_path'],'markdown':e['markdown_file'],'ledger_contract_no':best[4]['合同号'],'ledger_contract_name':best[4]['合同名称'],'ledger_customer':best[4]['客户名称'],'amount_match':best[1],'party_a_match':best[2],'contract_no_hint':best[3],'score':round(best[0],3),'second_score':round(second[0],3),'status':status})
with OUT.open('w',encoding='utf-8-sig',newline='') as f:
 w=csv.DictWriter(f,fieldnames=report[0].keys()); w.writeheader();w.writerows(report)
print('sources=',len(report),'confirmed=',sum(x['status'].startswith('confirmed') for x in report),'review=',sum(x['status']=='needs_review' for x in report));print(OUT)
