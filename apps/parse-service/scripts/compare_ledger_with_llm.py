#!/usr/bin/env python3
"""用 Qwen 从 59 个合同候选包提取身份字段，并与审核台账对照（只读，不写数据库）。"""
from __future__ import annotations
import csv, json, pathlib, re, sys
from difflib import SequenceMatcher
import httpx
from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[3]
PDF_ROOT, MD_ROOT = ROOT / "data/pdf", ROOT / "data/md-pdf"
LEDGER = ROOT / "demo/合同台账-V2.xlsx"
OUT = MD_ROOT / "llm-ledger-comparison.csv"
CACHE = MD_ROOT / "llm-contract-identities.json"
BASE_URL, MODEL = "http://192.168.121.32:6013/v1", "Qwen3.8-27B"

def norm(v: object) -> str:
    raw = str(v or "").strip().lower()
    if raw in {"null", "none", "未知", "无"}:
        return ""
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]", "", raw)

def packages():
    manifest = json.loads((MD_ROOT / "manifest.json").read_text(encoding="utf-8"))
    groups = {}
    for item in manifest["entries"]:
        # 同内容 PDF 可共用一个 Markdown，但仍可能属于不同合同包；必须展开全部来源。
        for source_info in item["sources"]:
            source = pathlib.PurePosixPath(source_info["relative_path"])
            key = source.parts[0] if len(source.parts) > 1 else source.stem
            groups.setdefault(key, []).append((source, item))
    return groups

def primary(files):
    def score(p):
        n=p.name
        return (100 if any(x in n for x in ("框架协议","协议书","合同")) else 0) - (80 if any(x in n for x in ("技术协议","附件","报价","说明书")) else 0)
    return max(files, key=lambda x: score(x[0]))

def ledger_rows():
    ws=load_workbook(LEDGER, read_only=True, data_only=True)["合同台账"]
    headers=[str(c or "").strip() for c in next(ws.iter_rows(min_row=3,max_row=3,values_only=True))]
    return [{headers[i]: v for i,v in enumerate(row) if i < len(headers)} for row in ws.iter_rows(min_row=5,values_only=True) if any(v is not None for v in row)]

def extract(client, markdown):
    prompt = '''从合同 Markdown 提取身份字段。只返回 JSON，不要解释：
{"contract_no": "我方合同号或null", "contract_name":"合同名称或null", "customer_name":"甲方客户或null"}。
不要把采购编号、项目编号当合同号。\n\n''' + markdown[:180000]
    r=client.post("/chat/completions", json={"model":MODEL,"messages":[{"role":"user","content":prompt}],"temperature":0,"response_format":{"type":"json_object"}})
    r.raise_for_status()
    return json.loads(r.json()["choices"][0]["message"]["content"])

def score(identity,row,key):
    a=norm(identity.get("contract_no")); b=norm(row.get("合同号"))
    if a and a==b: return 1.0
    # 文件包常含我方合同号短码（如 RJ-2026020），而台账前缀可能是 HSKJ/C-。
    key_codes=re.findall(r"(?:cm|rj|qc|qt|dl|hj)\d{7}", norm(key), flags=re.I)
    if any(code.lower() in b for code in key_codes): return 0.99
    vals=[SequenceMatcher(None,norm(identity.get(k)),norm(row.get(col))).ratio() for k,col in (("contract_name","合同名称"),("customer_name","客户名称")) if identity.get(k) and row.get(col)]
    vals += [SequenceMatcher(None,norm(key),norm(row.get(col))).ratio() for col in ("合同号","合同名称")]
    return max(vals, default=0.0)

def main():
    groups, rows = packages(), ledger_rows()
    if len(groups)!=len(rows): raise SystemExit(f"数量不一致：合同包 {len(groups)}，台账 {len(rows)}")
    cache=json.loads(CACHE.read_text()) if CACHE.exists() else {}
    client=httpx.Client(base_url=BASE_URL, timeout=300)
    for i,(key,files) in enumerate(sorted(groups.items()),1):
        if key not in cache:
            source,item=primary(files)
            cache[key]={"primary_source":str(source), **extract(client,(MD_ROOT/item["markdown_file"]).read_text(encoding="utf-8"))}
            CACHE.write_text(json.dumps(cache,ensure_ascii=False,indent=2),encoding="utf-8")
        print(f"[{i}/{len(groups)}] {key}")
    unmatched=set(range(len(rows))); report=[]
    # 先锁定唯一的高置信合同号，防止贪心弱匹配占用它们的台账行。
    remaining=[]
    for key,identity in cache.items():
        exact=[idx for idx,r in enumerate(rows) if score(identity,r,key)>=.99]
        if len(exact)==1 and exact[0] in unmatched:
            idx=exact[0]; unmatched.remove(idx)
            row=rows[idx]
            report.append({"package_key":key,"primary_source":identity["primary_source"],"llm_contract_no":identity.get("contract_no"),"llm_contract_name":identity.get("contract_name"),"llm_customer_name":identity.get("customer_name"),"ledger_contract_no":row.get("合同号"),"ledger_contract_name":row.get("合同名称"),"ledger_customer_name":row.get("客户名称"),"match_score":"1.000","status":"matched"})
        else: remaining.append((key,identity))
    for key,identity in remaining:
        ranked=sorted(((score(identity,r,key),idx) for idx,r in enumerate(rows) if idx in unmatched),reverse=True)
        s,idx=ranked[0]; unmatched.remove(idx)
        row=rows[idx]
        report.append({"package_key":key,"primary_source":identity["primary_source"],"llm_contract_no":identity.get("contract_no"),"llm_contract_name":identity.get("contract_name"),"llm_customer_name":identity.get("customer_name"),"ledger_contract_no":row.get("合同号"),"ledger_contract_name":row.get("合同名称"),"ledger_customer_name":row.get("客户名称"),"match_score":f"{s:.3f}","status":"matched" if s>=0.72 else "needs_review"})
    with OUT.open("w",encoding="utf-8-sig",newline="") as f:
        w=csv.DictWriter(f,fieldnames=list(report[0])); w.writeheader(); w.writerows(report)
    print(f"报告：{OUT}；待人工复核：{sum(x['status']=='needs_review' for x in report)}/{len(report)}")
if __name__=='__main__': main()
