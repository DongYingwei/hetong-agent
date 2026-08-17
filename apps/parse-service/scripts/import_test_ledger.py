"""导入已审核的「合同台账-V2.xlsx」作为可回滚的结构化查询测试数据。

该工作簿是已审核台账，不是合同原文来源：本脚本只写 PostgreSQL 正式库
``contracts`` 与 ``contract_module_hits``，不会建立 Milvus 向量。因此它适合
Text-to-SQL/台账页测试，不能作为 RAG 原文问答的测试数据。

用法（在 apps/parse-service 下）：
    python3 scripts/import_test_ledger.py
    python3 scripts/import_test_ledger.py --purge

``--purge`` 仅删除本脚本以固定测试批次标识写入的记录；外键会清理对应的
PG module hits/chunks，绝不删除人工导入或 PDF 解析入库的合同。
"""

from __future__ import annotations

import argparse
import pathlib
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent / "src"))

import psycopg  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from jinguan_parse.config import load_settings  # noqa: E402
from jinguan_parse.taxonomy import load_matcher  # noqa: E402


REPO_ROOT = pathlib.Path(__file__).resolve().parents[3]
LEDGER_PATH = REPO_ROOT / "demo" / "合同台账-V2.xlsx"
SHEET_NAME = "合同台账"
BATCH_MARKER = "test-ledger-v2-import-20260814"

# Excel 第 3 行「台账明细」的中文表头 → contracts 列。
COLUMNS = {
    "合同号": "contract_no", "客户名称": "customer_name", "合同名称": "contract_name",
    "考核线": "assessment_line", "中标编号": "bid_no", "关联主合同号": "related_main_no",
    "框架简称": "framework_alias", "客方合同号": "customer_contract_no",
    "签约法人体": "signing_entity", "合同类型": "contract_type", "签约时间": "sign_date",
    "开始时间": "start_date", "结束时间": "end_date", "金额属性": "amount_type",
    "合同金额（含税）": "amount", "税率": "tax_rate", "结算条款": "settlement_terms",
    "是否涉及后评估": "post_eval", "履约保证金金额": "deposit_amount",
    "履约保证金退还条件": "deposit_refund", "仲裁方式": "arbitration",
    "授权人": "authorizer", "合同状态": "status",
}
MODULE_COLUMNS = {
    "服务内容": "service", "技术要求": "tech", "岗位说明": "role", "人员需求": "staff",
}


def clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return None if not text or text in {"/", "／", "-"} else text


def as_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if text is None:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "").replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    # 如“签订之日”是已审核台账中的合法约定式日期，但不是可物化的日历日期。
    # contracts.sign_date 是 DATE，不能伪造；以 NULL 保留“不可按日期统计”的语义。
    return None


def as_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = clean(value)
    if text is None:
        return None
    text = text.replace(",", "").replace("元", "").replace("人民币", "")
    multiplier = Decimal(1)
    if text.endswith("万"):
        text = text[:-1]
        multiplier = Decimal("10000")
    elif text.endswith("亿"):
        text = text[:-1]
        multiplier = Decimal("100000000")
    try:
        return Decimal(text) * multiplier
    except InvalidOperation as exc:
        raise ValueError(f"无法解析金额：{value!r}") from exc


def load_rows() -> list[dict[str, object]]:
    wb = load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        headers = [clean(v) for v in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
        index = {header: i for i, header in enumerate(headers) if header}
        required = set(COLUMNS) | set(MODULE_COLUMNS) | {"是否包含AI关键词"}
        missing = required - set(index)
        if missing:
            raise ValueError(f"台账缺少列：{sorted(missing)}")
        rows: list[dict[str, object]] = []
        for excel_row, values in enumerate(ws.iter_rows(min_row=5, values_only=True), 5):
            contract_no = clean(values[index["合同号"]])
            if not contract_no:
                continue
            item = {db_col: clean(values[index[xlsx_col]]) for xlsx_col, db_col in COLUMNS.items()}
            item["sign_date"] = as_date(values[index["签约时间"]])
            item["start_date"] = as_date(values[index["开始时间"]])
            item["end_date"] = as_date(values[index["结束时间"]])
            item["amount"] = as_decimal(values[index["合同金额（含税）"]])
            item["deposit_amount"] = as_decimal(values[index["履约保证金金额"]])
            item["module_texts"] = {
                key: clean(values[index[xlsx_col]]) for xlsx_col, key in MODULE_COLUMNS.items()
            }
            item["tag_ai"] = 1 if clean(values[index["是否包含AI关键词"]]) == "是" else 0
            item["excel_row"] = excel_row
            rows.append(item)
        return rows
    finally:
        wb.close()


def purge(conn: psycopg.Connection) -> int:
    with conn.cursor() as cur:
        cur.execute("DELETE FROM contracts WHERE confirmed_by = %s", (BATCH_MARKER,))
        deleted = cur.rowcount
    conn.commit()
    return deleted


def import_rows(conn: psycopg.Connection, rows: list[dict[str, object]]) -> int:
    matcher = load_matcher(str(LEDGER_PATH))
    with conn.cursor() as cur:
        cur.execute("SELECT module_key FROM contract_modules WHERE enabled ORDER BY sort_order")
        enabled_modules = {row[0] for row in cur.fetchall()}
        unknown = set(MODULE_COLUMNS.values()) - enabled_modules
        if unknown:
            raise RuntimeError(f"数据库未启用所需模块：{sorted(unknown)}")

        inserted = 0
        for row in rows:
            module_texts = row.pop("module_texts")
            excel_row = row.pop("excel_row")
            columns = list(COLUMNS.values()) + ["tag_ai", "confirmed", "confirmed_by", "confirmed_at"]
            values = [row[column] for column in COLUMNS.values()] + [row["tag_ai"], 1, BATCH_MARKER, datetime.now().astimezone()]
            placeholders = ", ".join(["%s"] * len(values))
            try:
                cur.execute(
                    f"INSERT INTO contracts ({', '.join(columns)}) VALUES ({placeholders}) RETURNING id",
                    values,
                )
            except psycopg.errors.UniqueViolation as exc:
                raise RuntimeError(f"Excel 第 {excel_row} 行合同号 {row['contract_no']!r} 已存在；请先 --purge 或处理冲突") from exc
            contract_id = cur.fetchone()[0]
            for module_key, raw_text in module_texts.items():
                matched = matcher.match(raw_text or "")
                cur.execute(
                    "INSERT INTO contract_module_hits "
                    "(contract_id, module_key, hit, keywords, category, raw_text, raw_text_ai_raw) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (contract_id, module_key, int(matched.hit), ",".join(matched.keywords) or None,
                     ",".join(matched.categories) or None, raw_text, raw_text),
                )
            inserted += 1
    conn.commit()
    return inserted


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--purge", action="store_true", help="清除本测试批次，不导入")
    args = parser.parse_args()
    if not LEDGER_PATH.is_file():
        raise FileNotFoundError(f"未找到台账：{LEDGER_PATH}")
    settings = load_settings(".env")
    with psycopg.connect(settings.pg_url) as conn:
        if args.purge:
            print(f"已清除测试批次合同：{purge(conn)} 条（{BATCH_MARKER}）")
        else:
            rows = load_rows()
            print(f"准备导入已审核台账：{len(rows)} 条，批次={BATCH_MARKER}")
            print(f"已导入正式库：{import_rows(conn, rows)} 条；未建向量（Excel 无合同原文）。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
