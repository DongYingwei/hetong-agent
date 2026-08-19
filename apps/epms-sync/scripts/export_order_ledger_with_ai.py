#!/usr/bin/env python3
"""为 EPMS 全量订单 Excel 追加 AI 命中列；只生成新文件，不修改数据库或源文件。"""
from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def norm_order_no(value: object) -> str:
    """订单附件目录会把 /、\\ 改为 _，与导入订单台账保持同一匹配口径。"""
    return re.sub(r"[\\\\/]", "_", str(value or "").strip())


def main() -> None:
    parser = argparse.ArgumentParser(description="导出带 AI 命中列的订单 Excel")
    parser.add_argument("--xlsx", required=True, type=Path, help="EPMS 全量订单 Excel（不会被修改）")
    parser.add_argument("--ai-results", required=True, type=Path, help="ai_keyword_results.json")
    parser.add_argument("--output", required=True, type=Path, help="新生成的 Excel 路径")
    parser.add_argument("--sheet", default=None, help="工作表名称；缺省时使用第一个工作表")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        raise SystemExit(f"找不到订单 Excel：{args.xlsx}")
    if not args.ai_results.is_file():
        raise SystemExit(f"找不到 AI 结果文件：{args.ai_results}")
    if args.output.resolve() == args.xlsx.resolve():
        raise SystemExit("--output 必须是新文件，不能覆盖原始 Excel")

    raw_results = json.loads(args.ai_results.read_text(encoding="utf-8"))
    if not isinstance(raw_results, dict):
        raise SystemExit("AI 结果必须是以订单编号为 key 的 JSON 对象")
    verdict_by_no = {
        norm_order_no(order_no): "是" if isinstance(result, dict) and result.get("verdict") == "是" else "否"
        for order_no, result in raw_results.items()
    }

    workbook = load_workbook(args.xlsx)
    sheet = workbook[args.sheet] if args.sheet else workbook.worksheets[0]
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value is not None}
    order_col = headers.get("订单编号")
    if not order_col:
        raise SystemExit(f"工作表 {sheet.title} 第 1 行未找到“订单编号”列")

    ai_col = headers.get("是否包含AI关键词")
    if ai_col is None:
        ai_col = sheet.max_column + 1
        source = sheet.cell(1, max(1, ai_col - 1))
        target = sheet.cell(1, ai_col, "是否包含AI关键词")
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        sheet.column_dimensions[target.column_letter].width = max(18, sheet.column_dimensions[source.column_letter].width or 0)

    total = matched = ai_count = 0
    for row_no in range(2, sheet.max_row + 1):
        order_no = sheet.cell(row_no, order_col).value
        if order_no is None or not str(order_no).strip():
            continue
        total += 1
        key = norm_order_no(order_no)
        verdict = verdict_by_no.get(key, "否")
        matched += int(key in verdict_by_no)
        ai_count += int(verdict == "是")
        cell = sheet.cell(row_no, ai_col, verdict)
        source = sheet.cell(row_no, max(1, ai_col - 1))
        cell._style = copy(source._style)
        cell.alignment = copy(source.alignment)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(json.dumps({
        "output": str(args.output),
        "orders": total,
        "matched_ai_results": matched,
        "ai_orders": ai_count,
        "unmatched_orders": total - matched,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
